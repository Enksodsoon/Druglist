#!/usr/bin/env python3
"""Import the organized guideline/regimen patch workbook as a safe runtime overlay."""
from __future__ import annotations

import json
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent))

from engine_common import ROOT, clean, ensure_dirs, now_iso, read_json, read_xlsx_sheet, stable_id, write_json, write_report

EXPECTED_WORKBOOK = ROOT / "source_workbooks/patches/druglist_patch_organized_merge_ready_lossless_20260516.xlsx"
DOWNLOAD_CANDIDATES = [
    Path("/Users/a12/Downloads/druglist_patch_organized_merge_ready_lossless_20260516.xlsx"),
    Path("/Users/a12/Downloads/druglist_patch_organized_merge_ready_20260516.xlsx"),
]

IMPORT_SHEETS = {
    "COPY_Action_Summary": ("action_summary.json", ""),
    "COPY_Fast_Regimen_Master": ("fast_regimen_patch.json", "Regimen_ID"),
    "COPY_Complaint_Index": ("complaint_index_patch.json", "Map_ID"),
    "COPY_Peds_Dose_Shortcuts": ("peds_dose_shortcuts_patch.json", "Shortcut_ID"),
    "COPY_Drug_Master_Lookup": ("drug_master_lookup_patch.json", "Drug_Key"),
    "COPY_Drug_Rules": ("drug_rules_patch.json", "Rule_ID"),
    "COPY_Manual_Review": ("manual_review_patch.json", ""),
    "COPY_Guideline_Dose_Ref": ("guideline_dose_ref_patch.json", ""),
    "COPY_Evidence_Map": ("evidence_map_patch.json", ""),
}

MANUAL_BDS = {"", "NOT_IN_WORKBOOK", "NO_BDS", "BDS_REVIEW", "MANUAL"}
NON_DRUG_TERMS = [
    "referral",
    "refer",
    "imaging",
    "x-ray",
    "xray",
    "sleep service",
    "surgery",
    "cbt",
    "cpap",
    "dxa",
    "ercp",
    "cholecystectomy",
    "assessment",
    "advice",
    "procedure",
]
ACTIVE_SAFETY_TYPES = {"avoid", "block", "red_flag", "red_flag_gate", "er_gate", "referral_gate", "specialist_lock"}
ACTIVE_ABX_TYPES = {"antibiotic_gate", "stewardship", "antibiotic_stewardship"}
ACTIVE_GUIDANCE_TYPES = {"diagnostic", "core", "non-drug", "non_drug", "clinical_guidance"}
REQUIRED_DISEASE_KEYS = {
    "allergic_rhinitis",
    "epistaxis",
    "bronchiolitis_child_assessment",
    "secondary_infected_eczema",
    "low_back_pain_over16",
    "sciatica_over16",
    "suspected_rheumatoid_arthritis",
    "suspected_osahs_over16",
    "insect_bite_sting_local_reaction",
    "tinnitus_assessment_referral",
    "impetigo_localised_nonbullous",
    "dyspepsia_uninvestigated_adult",
    "gord_adult_empiric",
    "cellulitis_erysipelas_nonsevere_adult",
    "acute_cough_no_abx",
    "acute_pyelonephritis_adult",
    "lower_uti_nonpregnant_woman",
    "acute_sore_throat_no_abx",
    "acute_sinusitis_no_abx",
    "psoriasis_mild_topical_adult",
    "vitamin_d_prevention_specific_groups",
    "constipation_child_idiopathic_assessment",
    "ibs_adult_diagnosis_assessment",
    "osteoporosis_fracture_risk_assessment",
    "male_luts_initial_assessment",
    "gallstone_suspected_diagnosis",
}


def workbook_path() -> Path:
    env_path = os.environ.get("GUIDELINE_PATCH_WORKBOOK")
    candidates = [Path(env_path)] if env_path else []
    candidates.extend([EXPECTED_WORKBOOK, *DOWNLOAD_CANDIDATES])
    for path in candidates:
        if path and path.exists():
            return path
    raise FileNotFoundError(
        "Guideline patch workbook not found. Place it at "
        f"{EXPECTED_WORKBOOK} or set GUIDELINE_PATCH_WORKBOOK to the workbook path."
    )


def sheet_names(path: Path) -> list[str]:
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(path, read_only=True, data_only=True)
        return list(wb.sheetnames)
    except ModuleNotFoundError:
        from zipfile import ZipFile
        import xml.etree.ElementTree as ET

        ns = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
        with ZipFile(path) as zf:
            root = ET.fromstring(zf.read("xl/workbook.xml"))
            return [sheet.attrib["name"] for sheet in root.findall("a:sheets/a:sheet", ns)]


def raw_sheet_rows(path: Path, sheet: str) -> list[list[str]]:
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        return [[clean(cell) for cell in row] for row in ws.iter_rows(values_only=True)]
    except ModuleNotFoundError:
        return read_xlsx_sheet(path, sheet)


def norm_col(value: str) -> str:
    text = clean(value).replace("\ufeff", "")
    text = re.sub(r"\s+", "_", text)
    return text.strip("_")


def is_repeated_header(row: dict[str, str]) -> bool:
    matches = 0
    filled = 0
    for key, value in row.items():
        if value:
            filled += 1
            if norm_col(value).lower() == key.lower():
                matches += 1
    return filled > 0 and matches / filled >= 0.6


def records_for_sheet(path: Path, sheet: str, skipped: list[dict[str, Any]]) -> list[dict[str, str]]:
    rows = raw_sheet_rows(path, sheet)
    header_idx = next((idx for idx, row in enumerate(rows) if any(clean(cell) for cell in row)), None)
    if header_idx is None:
        skipped.append({"sheet": sheet, "row_number": None, "reason": "empty_sheet"})
        return []
    raw_headers = rows[header_idx]
    seen: dict[str, int] = {}
    headers = []
    for idx, header in enumerate(raw_headers, start=1):
        name = norm_col(header) or f"col_{idx}"
        seen[name] = seen.get(name, 0) + 1
        headers.append(name if seen[name] == 1 else f"{name}.{seen[name] - 1}")
    out = []
    for offset, row in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
        if not any(clean(cell) for cell in row):
            continue
        record = {headers[i]: clean(row[i]) if i < len(row) else "" for i in range(len(headers))}
        record["_sheet"] = sheet
        record["_workbook_row"] = str(offset)
        if is_repeated_header(record):
            skipped.append({"sheet": sheet, "row_number": offset, "reason": "repeated_header_row"})
            continue
        out.append(record)
    return out


def merge_latest_non_empty(rows: list[dict[str, str]]) -> dict[str, str]:
    merged = dict(rows[0])
    for row in rows[1:]:
        for key, value in row.items():
            if value:
                merged[key] = value
    return merged


def dedupe(sheet: str, rows: list[dict[str, str]], id_col: str, duplicates: list[dict[str, Any]], skipped: list[dict[str, Any]]) -> list[dict[str, str]]:
    if not id_col:
        return rows
    grouped: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        row_id = row.get(id_col, "")
        if not row_id:
            if sheet == "COPY_Drug_Master_Lookup":
                row_id = stable_id("PATCH_DRUG_ROW", f"{row.get('BDS_Code') or row.get('BDS')} {row.get('Display_Name') or row.get('Product')} {row.get('_workbook_row')}")
                row[id_col] = row_id
                row["_Import_Note"] = f"Generated stable import key because {id_col} was blank in workbook row."
            else:
                skipped.append({"sheet": sheet, "row_number": row.get("_workbook_row"), "reason": f"missing_{id_col}"})
                continue
        grouped[row_id].append(row)
    out = []
    for row_id, group in grouped.items():
        if len(group) > 1:
            duplicates.append({"sheet": sheet, "id_column": id_col, "id": row_id, "rows": group})
        out.append(merge_latest_non_empty(group))
    return sorted(out, key=lambda row: row.get(id_col, ""))


def yes(value: Any) -> bool:
    return clean(value).upper() in {"Y", "YES", "TRUE", "1", "ACTIVE", "ENABLED"}


def maybe_int(value: Any, default: int = 5) -> int:
    try:
        return int(float(clean(value)))
    except ValueError:
        return default


def non_drug_action(display: str, bds: str, order: str) -> bool:
    text = f"{display} {bds} {order}".lower()
    return any(term in text for term in NON_DRUG_TERMS)


def source_ids_from(row: dict[str, str]) -> list[str]:
    values = [row.get("Source_File", ""), row.get("Source_Anchor", ""), row.get("Evidence_ID", "")]
    ids = [stable_id("SRC", value) for value in values if clean(value)]
    return sorted(set(ids))


def product_maps() -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    products = read_json("data/core/drug_master_rebuilt.json", {"products": []}).get("products", [])
    by_bds = {clean(p.get("bds_code")): p for p in products if clean(p.get("bds_code"))}
    by_display = {clean(p.get("display_name")).lower(): p for p in products if clean(p.get("display_name"))}
    return by_bds, by_display


def peds_shortcut_verified(row: dict[str, str]) -> bool:
    return False


def build_line(row: dict[str, str], slot: str, disease_id: str, product_by_bds: dict[str, dict[str, Any]], product_by_display: dict[str, dict[str, Any]], peds_verified: set[tuple[str, str]]) -> tuple[dict[str, Any] | None, str | None]:
    display = row.get(f"{slot}_Display_Name", "")
    if not display:
        return None, f"blank_{slot}_Display_Name"
    bds = row.get(f"{slot}_BDS", "")
    order = row.get(f"{slot}_Order_Line", "")
    product = product_by_bds.get(bds) or product_by_display.get(display.lower()) or {}
    manual_bds = bds.upper() in MANUAL_BDS
    is_non_drug = non_drug_action(display, bds, order)
    missing = []
    readiness = "manual_review_required"
    source_status = "pending_manual_review"
    fast_allowed = False
    if is_non_drug:
        readiness = "non_drug_action"
        source_status = "not_applicable"
    if manual_bds and not is_non_drug:
        missing.append("workbook-supported BDS")
    if not product and bds and bds.upper() not in MANUAL_BDS:
        missing.append("matched product record")
    abx_flag = clean(row.get("Antibiotic_Indicated")).upper()
    if abx_flag in {"Y", "YES", "CONDITIONAL"}:
        missing.append("antibiotic rule criteria")
    if yes(row.get("Needs_Peds_Calc")) and (disease_id, slot) not in peds_verified:
        missing.append("verified exact pediatric shortcut")
    if not source_ids_from(row):
        missing.append("source evidence link")
    clinical_readiness = "manual_review_required" if readiness == "manual_review_required" else "manual_review_required"
    if is_non_drug:
        clinical_readiness = "manual_review_required"
    return {
        "line_id": stable_id("PATCH_LINE", f"{row.get('Regimen_ID')} {slot}"),
        "line_type": "RX_NOW" if slot.startswith("RX") else "SWAP",
        "slot": slot,
        "product_id": clean(product.get("id")) or (bds if bds and bds.upper() not in MANUAL_BDS else ""),
        "display_name": display,
        "bds": bds,
        "order_text": order,
        "duration_label": row.get(f"{slot}_Duration", ""),
        "dispense_label": row.get(f"{slot}_Dispense", ""),
        "pack_label": row.get(f"{slot}_Dispense", ""),
        "quick_side_effects": row.get(f"{slot}_Quick_SE", ""),
        "quick_caution": row.get(f"{slot}_Quick_Caution", ""),
        "source_ids": source_ids_from(row),
        "source_status": source_status,
        "clinical_readiness": clinical_readiness,
        "manual_review": True,
        "manual_review_required": True,
        "fast_mode_allowed": bool(fast_allowed),
        "missing_requirements": sorted(set(missing or ["manual clinical review"])),
        "clinical_audit_status": "manual_review_required",
        "regimen_safety_status": "warning",
        "pediatric_gate_status": "blocked" if yes(row.get("Needs_Peds_Calc")) else "not_applicable",
        "antibiotic_gate_status": "blocked" if abx_flag in {"Y", "YES", "CONDITIONAL"} else "not_applicable",
        "product_match_status": "matched" if product else "missing_product_match",
        "blocked_reason": "; ".join(sorted(set(missing))) if missing else "manual clinical review required",
        "next_action": "manual review before Fast RX NOW",
        "non_drug_action": is_non_drug,
    }, None


def transform_regimens(rows: list[dict[str, str]], skipped: list[dict[str, Any]]) -> list[dict[str, Any]]:
    product_by_bds, product_by_display = product_maps()
    peds_rows = read_json("data/imported_guideline_patch/peds_dose_shortcuts_patch.json", {"items": []}).get("items", [])
    peds_verified = {(row.get("Disease_Key", ""), row.get("Drug_Slot", "")) for row in peds_rows if row.get("verified_usable")}
    out = []
    for row in rows:
        lines = []
        for slot in [f"RX{i}" for i in range(1, 5)] + [f"SWAP{i}" for i in range(1, 4)]:
            line, reason = build_line(row, slot, row.get("Disease_Key", ""), product_by_bds, product_by_display, peds_verified)
            if reason:
                skipped.append({"sheet": "COPY_Fast_Regimen_Master", "row_number": row.get("_workbook_row"), "regimen_id": row.get("Regimen_ID"), "reason": reason})
            if line:
                lines.append(line)
        out.append(
            {
                "regimen_id": row.get("Regimen_ID"),
                "disease_id": row.get("Disease_Key"),
                "display_name": row.get("Disease_Name") or row.get("Disease_Key"),
                "disease_name": row.get("Disease_Name"),
                "complaint_group": row.get("Complaint_Group"),
                "age_group": row.get("Age_Group"),
                "severity": row.get("Severity"),
                "is_default": yes(row.get("Default_Row")),
                "display_order": maybe_int(row.get("Display_Order")),
                "icd10_primary": row.get("ICD10_Primary"),
                "icd10_secondary": row.get("ICD10_Secondary"),
                "likelihood_label": row.get("Likelihood_Label"),
                "key_clue": row.get("Key_Clue"),
                "use_when": row.get("Use_When"),
                "avoid_when": row.get("Avoid_When"),
                "antibiotic_indicated": row.get("Antibiotic_Indicated"),
                "needs_peds_calc": row.get("Needs_Peds_Calc"),
                "red_flags_one_line": row.get("Red_Flags_One_Line"),
                "notes_internal": row.get("Notes_Internal"),
                "verification_status": row.get("Verification_Status"),
                "manual_review_note": row.get("Manual_Review_Note"),
                "source_file": row.get("Source_File"),
                "source_anchor": row.get("Source_Anchor"),
                "source_ids": source_ids_from(row),
                "source_status": "pending_manual_review",
                "manual_review": True,
                "manual_review_required": True,
                "fast_mode_allowed": False,
                "import_source": "guideline_patch_20260516",
                "lines": lines,
            }
        )
    return out


def transform_complaints(rows: list[dict[str, str]], skipped: list[dict[str, Any]]) -> list[dict[str, Any]]:
    out = []
    for row in rows:
        if not yes(row.get("Enabled")):
            skipped.append({"sheet": "COPY_Complaint_Index", "row_number": row.get("_workbook_row"), "map_id": row.get("Map_ID"), "reason": "Enabled_not_Y_preserved_in_import_only"})
            continue
        complaint = row.get("Input_Phrase") or row.get("Normalized_Input")
        if not complaint or not row.get("Disease_Key"):
            skipped.append({"sheet": "COPY_Complaint_Index", "row_number": row.get("_workbook_row"), "map_id": row.get("Map_ID"), "reason": "missing_complaint_or_disease_key"})
            continue
        out.append(
            {
                "complaint_id": row.get("Map_ID"),
                "complaint": complaint,
                "normalized_input": row.get("Normalized_Input") or complaint.lower(),
                "complaint_group": row.get("Complaint_Group"),
                "disease_id": row.get("Disease_Key"),
                "age_group": row.get("Age_Group"),
                "priority": maybe_int(row.get("Priority")),
                "match_type": row.get("Match_Type") or "alias",
                "enabled": True,
                "notes": row.get("Notes"),
                "source_status": "pending_manual_review",
                "manual_review": True,
                "import_source": "guideline_patch_20260516",
            }
        )
    return out


def transform_rules(rows: list[dict[str, str]]) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    validation, antibiotic = [], []
    for row in rows:
        rule_type = clean(row.get("Rule_Type")).lower()
        category = "manual_review"
        active = False
        if any(token in rule_type for token in ACTIVE_SAFETY_TYPES):
            category, active = "guideline_safety", True
        if any(token in rule_type for token in ACTIVE_ABX_TYPES):
            category, active = "antibiotic_stewardship", True
        if rule_type in ACTIVE_GUIDANCE_TYPES:
            category, active = "clinical_guidance", True
        if any(token in rule_type for token in ["product", "dose", "uncertain"]):
            category, active = "manual_review", False
        payload = {
            "rule_id": row.get("Rule_ID"),
            "title": row.get("Action_Message")[:120] or row.get("Rule_ID"),
            "disease_id": row.get("Disease_Key"),
            "rule_type": row.get("Rule_Type"),
            "category": category,
            "target": row.get("Target"),
            "condition": row.get("Condition"),
            "message": row.get("Action_Message"),
            "action_message": row.get("Action_Message"),
            "active": active,
            "severity": "review",
            "source_ids": source_ids_from(row),
            "source_status": "pending_manual_review",
            "manual_review": True,
            "import_source": "guideline_patch_20260516",
        }
        (antibiotic if category == "antibiotic_stewardship" else validation).append(payload)
    return validation, antibiotic


def overlay_runtime(imported: dict[str, list[dict[str, Any]]], report_counts: dict[str, Any]) -> None:
    ensure_dirs("data/core", "data/safety", "data/pediatric", "data/meta")
    diseases = read_json("data/core/disease_master.json", {"meta": {}, "diseases": []})
    disease_by_id = {row.get("disease_id"): row for row in diseases.get("diseases", [])}
    for regimen in imported["regimens"]:
        disease_id = regimen["disease_id"]
        current = disease_by_id.get(disease_id, {})
        aliases = set(current.get("aliases") or [])
        aliases.update([regimen.get("disease_name") or disease_id])
        disease_by_id[disease_id] = {
            **current,
            "disease_id": disease_id,
            "display_name": regimen.get("disease_name") or current.get("display_name") or disease_id.replace("_", " ").title(),
            "aliases": sorted(a for a in aliases if a),
            "category": regimen.get("complaint_group") or current.get("category", ""),
            "icd10_primary": regimen.get("icd10_primary", ""),
            "source_ids": sorted(set((current.get("source_ids") or []) + (regimen.get("source_ids") or []))),
            "source_status": "pending_manual_review",
            "manual_review": True,
            "import_source": "guideline_patch_20260516",
        }
    for complaint in imported["complaints"]:
        disease_id = complaint["disease_id"]
        if disease_id not in disease_by_id:
            disease_by_id[disease_id] = {
                "disease_id": disease_id,
                "display_name": disease_id.replace("_", " ").title(),
                "aliases": [complaint["complaint"]],
                "category": complaint.get("complaint_group", ""),
                "source_ids": [],
                "source_status": "pending_manual_review",
                "manual_review": True,
                "import_source": "guideline_patch_20260516",
            }
        else:
            aliases = set(disease_by_id[disease_id].get("aliases") or [])
            aliases.add(complaint["complaint"])
            disease_by_id[disease_id]["aliases"] = sorted(aliases)
    diseases["diseases"] = sorted(disease_by_id.values(), key=lambda row: row.get("disease_id", ""))
    diseases["meta"] = {**diseases.get("meta", {}), "guideline_patch_disease_count": len(imported["regimens"])}
    write_json("data/core/disease_master.json", diseases)

    complaints = read_json("data/core/complaint_index.json", {"meta": {}, "items": []})
    by_id = {row.get("complaint_id"): row for row in complaints.get("items", [])}
    by_id.update({row["complaint_id"]: row for row in imported["complaints"]})
    complaints["items"] = sorted(by_id.values(), key=lambda row: (row.get("disease_id", ""), maybe_int(row.get("priority"))))
    complaints["meta"] = {**complaints.get("meta", {}), "guideline_patch_complaint_count": len(imported["complaints"])}
    write_json("data/core/complaint_index.json", complaints)

    regimens = read_json("data/core/fast_regimen_master.json", {"meta": {}, "regimens": []})
    by_regimen = {row.get("regimen_id"): row for row in regimens.get("regimens", [])}
    by_regimen.update({row["regimen_id"]: row for row in imported["regimens"]})
    regimens["regimens"] = sorted(by_regimen.values(), key=lambda row: row.get("regimen_id", ""))
    regimens["meta"] = {**regimens.get("meta", {}), "guideline_patch_regimen_count": len(imported["regimens"])}
    write_json("data/core/fast_regimen_master.json", regimens)

    opd = read_json("data/core/opd_fast_index.json", {"meta": {}, "index": [], "layer_links": {}})
    complaint_items = complaints["items"]
    regimen_items = regimens["regimens"]
    opd["index"] = [
        {
            "disease_id": disease["disease_id"],
            "display_name": disease.get("display_name", disease["disease_id"]),
            "complaints": [c for c in complaint_items if c.get("disease_id") == disease["disease_id"]],
            "regimen_ids": [r["regimen_id"] for r in regimen_items if r.get("disease_id") == disease["disease_id"]],
            "manual_review": True,
        }
        for disease in diseases["diseases"]
    ]
    opd["meta"] = {**opd.get("meta", {}), "guideline_patch_overlay": True, "guideline_patch_regimen_count": len(imported["regimens"])}
    write_json("data/core/opd_fast_index.json", opd)

    validation, antibiotic = transform_rules(imported["raw_rules"])
    validation_payload = read_json("data/safety/validation_rules.json", {"meta": {}, "rules": []})
    antibiotic_payload = read_json("data/safety/antibiotic_stewardship.json", {"meta": {}, "rules": []})
    validation_by_id = {row.get("rule_id"): row for row in validation_payload.get("rules", []) if row.get("import_source") != "guideline_patch_20260516"}
    validation_by_id.update({row["rule_id"]: row for row in validation if row.get("rule_id")})
    antibiotic_by_id = {row.get("rule_id"): row for row in antibiotic_payload.get("rules", []) if row.get("import_source") != "guideline_patch_20260516"}
    antibiotic_by_id.update({row["rule_id"]: row for row in antibiotic if row.get("rule_id")})
    validation_payload["rules"] = list(validation_by_id.values())
    antibiotic_payload["rules"] = list(antibiotic_by_id.values())
    validation_payload["meta"] = {**validation_payload.get("meta", {}), "guideline_patch_rule_count": len(validation)}
    antibiotic_payload["meta"] = {**antibiotic_payload.get("meta", {}), "guideline_patch_rule_count": len(antibiotic)}
    write_json("data/safety/validation_rules.json", validation_payload)
    write_json("data/safety/antibiotic_stewardship.json", antibiotic_payload)

    write_json(
        "data/pediatric/imported_guideline_peds_shortcuts.json",
        {
            "meta": {"generated_at": now_iso(), "manual_review": True, "shortcut_count": len(imported["peds_outputs"])},
            "items": imported["peds_outputs"],
        },
    )

    write_json(
        "data/meta/guideline_patch_manual_review_queue.json",
        {
            "meta": {"generated_at": now_iso(), "manual_review": True, "item_count": len(imported["manual_review"])},
            "items": imported["manual_review"],
        },
    )
    write_json("data/meta/guideline_patch_runtime_summary.json", report_counts)


def main() -> int:
    path = workbook_path()
    ensure_dirs("data/imported_guideline_patch", "reports")
    skipped: list[dict[str, Any]] = []
    duplicates: list[dict[str, Any]] = []
    sheet_list = sheet_names(path)
    raw_audit = [name for name in sheet_list if name.startswith("RAW_")]
    normalized: dict[str, list[dict[str, str]]] = {}
    for sheet, (filename, id_col) in IMPORT_SHEETS.items():
        if sheet not in sheet_list:
            skipped.append({"sheet": sheet, "row_number": None, "reason": "missing_expected_copy_sheet"})
            normalized[sheet] = []
            continue
        rows = records_for_sheet(path, sheet, skipped)
        normalized[sheet] = dedupe(sheet, rows, id_col, duplicates, skipped)
        write_json(f"data/imported_guideline_patch/{filename}", {"meta": {"generated_at": now_iso(), "sheet": sheet, "input_workbook": str(path), "row_count": len(normalized[sheet])}, "items": normalized[sheet]})

    peds_rows = []
    for row in normalized["COPY_Peds_Dose_Shortcuts"]:
        row["verified_usable"] = peds_shortcut_verified(row)
        peds_rows.append(
            {
                "shortcut_id": row.get("Shortcut_ID"),
                "product_id": row.get("BDS") if row.get("BDS", "").upper() not in MANUAL_BDS else "",
                "display_name": row.get("Display_Name"),
                "disease_id": row.get("Disease_Key"),
                "drug_slot": row.get("Drug_Slot"),
                "dose_output_status": "verified_shortcut" if row["verified_usable"] else "manual_review",
                "auto_dose_enabled": bool(row["verified_usable"]),
                "source_ids": source_ids_from(row),
                "review_reasons": [] if row["verified_usable"] else ["missing exact product/dose/age-weight/frequency/duration/source requirement"],
                "import_source": "guideline_patch_20260516",
            }
        )
    write_json("data/imported_guideline_patch/peds_dose_shortcuts_patch.json", {"meta": {"generated_at": now_iso(), "sheet": "COPY_Peds_Dose_Shortcuts", "input_workbook": str(path), "row_count": len(normalized["COPY_Peds_Dose_Shortcuts"])}, "items": normalized["COPY_Peds_Dose_Shortcuts"]})

    regimens = transform_regimens(normalized["COPY_Fast_Regimen_Master"], skipped)
    complaints = transform_complaints(normalized["COPY_Complaint_Index"], skipped)
    manual_review = [
        {
            "review_id": stable_id("PATCH_REVIEW", f"{row.get('Issue')} {row.get('Affected_Key')} {row.get('_workbook_row')}"),
            "issue": row.get("Issue"),
            "severity": row.get("Severity"),
            "affected_key": row.get("Affected_Key"),
            "why_it_matters": row.get("Why_It_Matters"),
            "recommended_action": row.get("Recommended_Action"),
            "source_status": "pending_manual_review",
            "manual_review": True,
            "import_source": "guideline_patch_20260516",
        }
        for row in normalized["COPY_Manual_Review"]
    ]
    imported = {
        "regimens": regimens,
        "complaints": complaints,
        "peds_outputs": peds_rows,
        "manual_review": manual_review,
        "raw_rules": normalized["COPY_Drug_Rules"],
    }
    line_rows = [line for regimen in regimens for line in regimen["lines"]]
    report_counts = {
        "generated_at": now_iso(),
        "input_workbook": str(path),
        "copy_sheet_counts": {sheet: len(rows) for sheet, rows in normalized.items()},
        "raw_audit_sheets": raw_audit,
        "skipped_rows": len(skipped),
        "duplicate_groups": len(duplicates),
        "disease_key_count": len({r.get("Disease_Key") for rows in normalized.values() for r in rows if r.get("Disease_Key")}),
        "active_rows": sum(1 for line in line_rows if line.get("fast_mode_allowed")),
        "manual_review_required_rows": sum(1 for line in line_rows if line.get("manual_review_required")),
        "blocked_non_drug_rows": sum(1 for line in line_rows if line.get("non_drug_action")),
        "not_in_workbook_rows": sum(1 for line in line_rows if line.get("bds") == "NOT_IN_WORKBOOK"),
        "bds_review_rows": sum(1 for line in line_rows if line.get("bds") == "BDS_REVIEW"),
        "regimen_count": len(regimens),
        "complaint_count": len(complaints),
        "manual_review_count": len(manual_review),
    }
    write_json("data/imported_guideline_patch/duplicate_rows.json", {"meta": {"generated_at": now_iso(), "count": len(duplicates)}, "items": duplicates})
    write_json("data/imported_guideline_patch/skipped_rows.json", {"meta": {"generated_at": now_iso(), "count": len(skipped)}, "items": skipped})
    write_json("data/imported_guideline_patch/import_manifest.json", report_counts)
    overlay_runtime(imported, report_counts)

    write_report(
        "reports/guideline_patch_import_report.md",
        "Guideline Patch Import Report",
        [
            ("Input Workbook", str(path)),
            ("Imported Row Counts", "\n".join(f"- {sheet}: {count}" for sheet, count in sorted(report_counts["copy_sheet_counts"].items()))),
            ("Skipped Rows", f"Skipped rows: {len(skipped)}\n\nSee `data/imported_guideline_patch/skipped_rows.json`."),
            ("Duplicate IDs", f"Duplicate ID groups: {len(duplicates)}\n\nSee `data/imported_guideline_patch/duplicate_rows.json`."),
            ("Disease Keys", f"Disease key count: {report_counts['disease_key_count']}"),
            ("Safety Counts", "\n".join(f"- {key}: {report_counts[key]}" for key in ["active_rows", "manual_review_required_rows", "blocked_non_drug_rows", "not_in_workbook_rows", "bds_review_rows"])),
        ],
    )
    write_report(
        "reports/guideline_patch_duplicate_report.md",
        "Guideline Patch Duplicate Report",
        [("Summary", f"Duplicate ID groups: {len(duplicates)}"), ("Details", "\n".join(f"- {d['sheet']} {d['id_column']}={d['id']}: {len(d['rows'])} rows" for d in duplicates) or "No duplicate stable IDs found.")],
    )
    write_report(
        "reports/guideline_patch_safety_report.md",
        "Guideline Patch Safety Report",
        [
            ("Fast Mode", "Imported medication rows default to `fast_mode_allowed=false`; NOT_IN_WORKBOOK, BDS_REVIEW, pediatric-calculation, and antibiotic rows remain gated."),
            ("Counts", "\n".join(f"- {key}: {report_counts[key]}" for key in ["manual_review_required_rows", "not_in_workbook_rows", "bds_review_rows", "blocked_non_drug_rows"])),
        ],
    )
    disease_ids = {row["disease_id"] for row in regimens} | {row["disease_id"] for row in complaints}
    covered = sorted(key for key in REQUIRED_DISEASE_KEYS if key in disease_ids or any(key in item for item in disease_ids))
    missing = sorted(REQUIRED_DISEASE_KEYS - set(covered))
    write_report(
        "reports/guideline_patch_runtime_coverage.md",
        "Guideline Patch Runtime Coverage",
        [
            ("Runtime Overlay", f"Regimens: {len(regimens)}\n\nComplaints: {len(complaints)}"),
            ("Representative Disease Keys Covered", "\n".join(f"- {key}" for key in covered) or "None"),
            ("Representative Disease Keys Not Found Exactly", "\n".join(f"- {key}" for key in missing) or "None"),
        ],
    )
    print(json.dumps(report_counts, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
