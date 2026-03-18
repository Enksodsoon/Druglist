#!/usr/bin/env python3
"""Build app-ready seed JSON from workbook + current app seed.

- Reads `index.html` embedded seed as baseline.
- Reads workbook `Price_Estimates_Online` to update `pr` and metadata `price_updated_at`.
- Expands price coverage using generic/category/form fallback imputation.
- Applies pediatric parser upgrades for common concentration patterns.
- Adds explicit confidence metadata for price and pediatric quality.
- Writes build/app_seed.json and build/build_report.json.
"""
from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path
from statistics import median

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INDEX = ROOT / "index.html"
WB = ROOT / "source_workbooks" / "drug_list_final_userfriendly_engine_ready_v7.xlsx"
OUT = ROOT / "build" / "app_seed.json"
REPORT = ROOT / "build" / "build_report.json"


def norm_id(v: object) -> str:
    return re.sub(r"[^A-Z0-9]", "", str(v or "").upper())


def norm_generic(v: object) -> str:
    s = str(v or "").lower().strip()
    s = re.sub(r"\([^)]*\)", "", s)
    s = re.sub(r"[^a-z0-9+ ]", " ", s)
    return re.sub(r"\s+", " ", s).strip()



GENERIC_ALIASES = {
    "acetaminophen": "paracetamol",
    "paracetamol": "paracetamol",
    "amoxicillin clavulanate": "amoxicillin + clavulanic acid",
    "cetirizine dihydrochloride": "cetirizine",
    "chlorpheniramine maleate": "chlorpheniramine",
    "fexofenadine hcl": "fexofenadine",
    "xylometazoline hydrochloride": "xylometazoline",
}


def canonical_generic_key(v: object) -> str:
    g = norm_generic(v)
    return GENERIC_ALIASES.get(g, g)


def generic_keys(v: object) -> set[str]:
    g = canonical_generic_key(v)
    keys = {g} if g else set()
    if g:
        base = re.sub(r"\b(syrup|suspension|drops?|oral|solution|dry|junior|paediatric|pediatric)\b", "", g)
        base = re.sub(r"\s+", " ", base).strip()
        if base:
            keys.add(base)
    if '+' in g:
        keys.add(' + '.join(sorted(x.strip() for x in g.split('+'))))
    return {k for k in keys if k}

def parse_price(v: object) -> float | None:
    if isinstance(v, (int, float)):
        x = float(v)
        return x if x > 0 else None
    s = str(v or "").strip().replace(",", "")
    if not s:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)", s)
    if not m:
        return None
    x = float(m.group(1))
    return x if x > 0 else None


def clean_cell(v: object) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def row_cell(row: tuple[object, ...], idx: int) -> str:
    return clean_cell(row[idx]) if idx < len(row) else ""


def parse_concentration(drug: dict) -> tuple[float, str] | None:
    text = f"{drug.get('n','')} {drug.get('c','')}".lower().replace("μg", "mcg")
    text = text.replace("mg./", "mg/").replace("mcg./", "mcg/").replace("ml.", "ml")

    m = re.search(r"(\d+(?:\.\d+)?)\s*mg\s*/\s*(\d+(?:\.\d+)?)\s*ml", text, re.I)
    if m:
        mg, ml = float(m.group(1)), float(m.group(2))
        if ml > 0:
            return round(mg / ml, 4), m.group(0)

    m = re.search(r"(\d+(?:\.\d+)?)\s*%", text)
    if m and any(k in text for k in ["ml", "syrup", "solution", "drops", "susp"]):
        return round(float(m.group(1)) * 10, 4), m.group(0)

    m = re.search(r"(\d+(?:\.\d+)?)\s*mcg\s*/\s*(\d+(?:\.\d+)?)\s*ml", text, re.I)
    if m:
        mcg, ml = float(m.group(1)), float(m.group(2))
        if ml > 0:
            return round((mcg / 1000.0) / ml, 6), m.group(0)

    return None


def parse_unit_strength(drug: dict) -> tuple[float, str] | None:
    text = f"{drug.get('n','')} {drug.get('c','')}".lower().replace("μg", "mcg")

    m = re.search(r"(\d+(?:\.\d+)?)\s*mg(?!\s*/)", text, re.I)
    if m:
        return float(m.group(1)), "mg"

    m = re.search(r"(\d[\d,]*(?:\.\d+)?)\s*u(?:nit)?s?(?!\s*/)", text, re.I)
    if m:
        return float(m.group(1).replace(",", "")), "U"

    return None


def parse_ref_volume_ml(order_text: object, cv_text: object = "") -> tuple[float, float] | None:
    text = str(order_text or "").strip().lower()
    if not text:
        return None

    text = text.replace("–", "-").replace("—", "-")

    def parse_range(match: re.Match[str], factor: float) -> tuple[float, float]:
        lo = float(match.group(1))
        hi = float(match.group(2)) if match.group(2) else lo
        return lo * factor, hi * factor

    for pat, factor in [
        (r"(\d+(?:\.\d+)?)\s*(?:-\s*(\d+(?:\.\d+)?))?\s*tsp\b", 5.0),
        (r"(\d+(?:\.\d+)?)\s*(?:-\s*(\d+(?:\.\d+)?))?\s*teaspoons?\b", 5.0),
        (r"(\d+(?:\.\d+)?)\s*(?:-\s*(\d+(?:\.\d+)?))?\s*ml\b", 1.0),
        (r"(\d+(?:\.\d+)?)\s*(?:-\s*(\d+(?:\.\d+)?))?\s*dropper\b", 1.0),
    ]:
        m = re.search(pat, text, re.I)
        if m:
            return parse_range(m, factor)

    cv = str(cv_text or "").lower()
    m = re.search(r"mix\s*(\d+(?:\.\d+)?)\s*sachet\s*in\s*(\d+(?:\.\d+)?)\s*ml", cv, re.I)
    if m:
        return float(m.group(2)), float(m.group(2))

    return None


def format_decimal(v: float) -> str:
    if abs(v - round(v)) < 1e-9:
        return str(int(round(v)))
    return f"{v:.2f}".rstrip("0").rstrip(".")


def format_fraction_units(lo: float, hi: float, unit: str) -> str:
    def fmt(v: float) -> str:
        nearest_quarter = round(v * 4) / 4
        if abs(v - nearest_quarter) < 1e-6:
            v = nearest_quarter
        return format_decimal(v)

    left = fmt(lo)
    right = fmt(hi)
    span = left if abs(lo - hi) < 1e-9 else f"{left}–{right}"
    return f"{span} {unit}"


def apply_curated_fexofenadine_rule(drug: dict) -> bool:
    tl = drug.get("tl") or {}
    strength = parse_unit_strength(drug)
    if not strength:
        return False

    amt, unit = strength
    if unit != "mg":
        return False

    ag = drug.get("ag") or {}
    if abs(amt - 60.0) < 1e-6:
        tl.update({
            "s": "curated_generic_fallback",
            "rp": "Curated fexofenadine pediatric fallback",
            "rs": "0.5 tab BID",
            "rf": "BID",
            "rd": "5 days",
            "rdi": "1 box / unit as needed",
            "nt": "Curated pediatric fallback: 60 mg tablet mapped to 30 mg BID equivalent.",
        })
        ag["imin"] = max(int(ag.get("imin") or 0), 6)
    elif abs(amt - 180.0) < 1e-6:
        tl.update({
            "s": "curated_generic_fallback",
            "rp": "Curated fexofenadine adolescent fallback",
            "rs": "1 tab OD",
            "rf": "OD",
            "rd": "5 days",
            "rdi": "1 box / unit as needed",
            "nt": "Curated adolescent fallback: 180 mg tablet limited to age 12 years and above.",
        })
        ag["imin"] = max(int(ag.get("imin") or 0), 12)
        ag["it"] = "curated adolescent gate: fexofenadine 180 mg tablet"
    else:
        return False

    drug["tl"] = tl
    drug["ag"] = ag
    return True


def annotate_peds_quality(drug: dict) -> None:
    tl = drug.get("tl") or {}
    s = tl.get("s")
    nt = str(tl.get("nt") or "").lower()
    if s == "auto_mapped_from_same_generic_reference":
        tl["q"] = "auto_mapped"
    elif s == "reference_default_derived_from_same_generic":
        tl["q"] = "auto_mapped"
    elif s == "reference_default_filled_from_template":
        tl["q"] = "manual_unknown"
    elif s == "curated_generic_fallback":
        tl["q"] = "generic_uplift_manual"
    elif s == "calculator_ready_manual_target_needed":
        if "parser" in nt:
            tl["q"] = "parser_manual"
        elif "top pediatric generic uplift" in nt:
            tl["q"] = "generic_uplift_manual"
        else:
            tl["q"] = "manual_unknown"
    else:
        tl["q"] = "none"
    drug["tl"] = tl


def main() -> None:
    seed_match = re.search(r'<script id="seed" type="application/json">(.*?)</script>', INDEX.read_text(), re.S)
    if not seed_match:
        raise SystemExit("Cannot find seed JSON in index.html")
    seed = json.loads(seed_match.group(1))

    drugs = seed.get("dr", [])
    by_id = {norm_id(d.get("i")): d for d in drugs if d.get("i")}
    pd_templates = seed.get("pd", [])

    wb = load_workbook(WB, data_only=True, read_only=True)
    ws = wb["Price_Estimates_Online"]
    final_ws = wb["Final approved"] if "Final approved" in wb.sheetnames else wb["Final approved "]

    direct_price_updates = 0
    generic_fallback_updates = 0
    category_fallback_updates = 0
    form_fallback_updates = 0
    latest_date = None

    generic_price_pool: dict[str, list[float]] = {}
    category_price_pool: dict[str, list[float]] = {}
    form_price_pool: dict[str, list[float]] = {}

    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        bds = norm_id(r[0])
        generic = canonical_generic_key(r[2])
        price = parse_price(r[4])
        checked = r[7]

        if bds and price and bds in by_id:
            by_id[bds]["pr"] = float(price)
            by_id[bds]["price_source"] = "workbook_direct"
            by_id[bds]["price_confidence"] = "direct"
            direct_price_updates += 1

        if generic and price:
            for gk in generic_keys(generic):
                generic_price_pool.setdefault(gk, []).append(float(price))

        if checked:
            s = str(checked)
            if not latest_date or s > latest_date:
                latest_date = s

    for r in final_ws.iter_rows(min_row=3, values_only=True):
        if not r:
            continue
        bds = norm_id(r[6])
        if not bds or bds not in by_id:
            continue
        by_id[bds]["fa"] = {
            "subcategory": row_cell(r, 1),
            "product_code": row_cell(r, 2),
            "product_name": row_cell(r, 3),
            "pack_size": row_cell(r, 4),
            "composition": row_cell(r, 5),
            "medicine_code": row_cell(r, 6),
            "medicine_name": row_cell(r, 7),
            "instructions_th": row_cell(r, 8),
            "instructions_en": row_cell(r, 9),
            "online_pack_price_thb": row_cell(r, 10),
            "online_unit_price_thb": row_cell(r, 11),
            "price_source_url": row_cell(r, 12),
            "price_checked_date": row_cell(r, 13),
            "price_notes": row_cell(r, 14),
            "category": row_cell(r, 15),
        }

    for d in drugs:
        pr = d.get("pr")
        if not (isinstance(pr, (int, float)) and pr > 0):
            continue
        cc = str(d.get("cc") or "").strip().lower()
        fm = str(d.get("f") or "").strip().lower()
        if cc:
            category_price_pool.setdefault(cc, []).append(float(pr))
        if fm:
            form_price_pool.setdefault(fm, []).append(float(pr))

    for d in drugs:
        if isinstance(d.get("pr"), (int, float)) and d.get("pr") > 0:
            if d.get("price_source") == "workbook_direct":
                d["price_confidence"] = "direct"
            else:
                gvals_exist = []
                for gk in generic_keys(d.get("g")):
                    gvals_exist.extend(generic_price_pool.get(gk, []))
                if gvals_exist:
                    d["price_source"] = "generic_reclassified_existing"
                    d["price_confidence"] = "generic_imputed"
                else:
                    d.setdefault("price_source", "existing")
                    d.setdefault("price_confidence", "category_imputed")
            continue

        vals = []
        for gk in generic_keys(d.get("g")):
            vals.extend(generic_price_pool.get(gk, []))
        if vals:
            d["pr"] = float(round(median(vals), 2))
            d["price_source"] = "generic_median_imputed"
            d["price_confidence"] = "generic_imputed"
            generic_fallback_updates += 1
            continue

        cc = str(d.get("cc") or "").strip().lower()
        cv = category_price_pool.get(cc)
        if cv:
            d["pr"] = float(round(median(cv), 2))
            d["price_source"] = "category_median_imputed"
            d["price_confidence"] = "category_imputed"
            category_fallback_updates += 1
            continue

        fm = str(d.get("f") or "").strip().lower()
        fv = form_price_pool.get(fm)
        if fv:
            d["pr"] = float(round(median(fv), 2))
            d["price_source"] = "form_median_imputed"
            d["price_confidence"] = "form_imputed"
            form_fallback_updates += 1

    upgraded_parse = 0
    upgraded_top_generic = 0
    upgraded_reference_manual = 0
    upgraded_reference_fraction = 0
    upgraded_curated_generic = 0
    top_peds_generics = {
        "paracetamol", "ibuprofen", "cetirizine", "loratadine", "chlorpheniramine", "amoxicillin",
        "azithromycin", "salbutamol", "bromhexine", "simethicone", "nystatin", "racecadotril",
    }

    curated_peds_generics = {"paracetamol", "ibuprofen", "amoxicillin", "cetirizine", "salbutamol", "bromhexine"}

    pd_rows_by_bds: dict[str, list[dict]] = {}
    pd_rows_by_generic: dict[str, list[dict]] = {}
    for tpl in pd_templates:
        for row in tpl.get("r", []):
            bid = norm_id(row.get("b") or row.get("i"))
            if bid:
                pd_rows_by_bds.setdefault(bid, []).append(row)
                ref_drug = by_id.get(bid)
                if ref_drug:
                    for gk in generic_keys(ref_drug.get("g")):
                        pd_rows_by_generic.setdefault(gk, []).append(row)

    for d in drugs:
        tl = d.get("tl") or {}
        status = tl.get("s")
        if status in {"auto_mapped_from_same_generic_reference", "calculator_ready_manual_target_needed"}:
            annotate_peds_quality(d)
            continue

        form = str(d.get("f") or "").lower()
        has_liquid_hint = any(k in form for k in ["syrup", "susp", "solution", "drop", "spray", "liquid"])
        conc = parse_concentration(d)
        if conc and (has_liquid_hint or "ml" in str(d.get("c") or "").lower()):
            per_ml, raw = conc
            tl["s"] = "calculator_ready_manual_target_needed"
            tl["dm"] = tl.get("dm") or "mgkg"
            tl["dv"] = tl.get("dv") or 10
            tl["pc"] = {"kind": "mg", "per_ml": per_ml, "raw": raw}
            tl["nt"] = "Auto-upgraded by parser: concentration pattern recognized."
            d["tl"] = tl
            annotate_peds_quality(d)
            upgraded_parse += 1
            continue

        g = norm_generic(d.get("g"))
        if any(x in g for x in top_peds_generics) and status == "no_pediatric_target_found":
            tl["s"] = "calculator_ready_manual_target_needed"
            tl["dm"] = tl.get("dm") or "mgkg"
            tl["dv"] = tl.get("dv") or 10
            tl["nt"] = "Top pediatric generic uplift: manual target required for safe dosing confirmation."
            d["tl"] = tl
            annotate_peds_quality(d)
            upgraded_top_generic += 1
            continue

        if canonical_generic_key(d.get("g")) == "fexofenadine" and status == "no_pediatric_target_found":
            if apply_curated_fexofenadine_rule(d):
                annotate_peds_quality(d)
                upgraded_curated_generic += 1
                continue

        bid = norm_id(d.get("i"))
        direct_rows = pd_rows_by_bds.get(bid, [])
        if direct_rows and not any(tl.get(k) for k in ("rs", "rf", "rd", "rdi")):
            row = direct_rows[0]
            tl["rp"] = row.get("n") or d.get("n") or ""
            tl["rs"] = str(row.get("o") or "").strip()
            tl["rf"] = str(row.get("f") or "").strip()
            tl["rd"] = str(row.get("u") or "").strip()
            tl["rdi"] = str(row.get("p") or "").strip()
            if status != "no_pediatric_target_found":
                tl["s"] = "reference_default_filled_from_template"
                tl["nt"] = "Direct pediatric template default copied into the drug library fallback."
                d["tl"] = tl
                annotate_peds_quality(d)
                upgraded_reference_manual += 1
                continue

        form = str(d.get("f") or "").lower()
        is_solid_oral = any(k in form for k in ["tablet", "capsule", "cap", "tabs"])
        strength = parse_unit_strength(d)
        if status == "reference_exists_but_not_parseable" and is_solid_oral and strength:
            unit_strength, unit_kind = strength
            derived = None
            for gk in generic_keys(d.get("g")):
                for row in pd_rows_by_generic.get(gk, []):
                    ref_bid = norm_id(row.get("b") or row.get("i"))
                    ref_drug = by_id.get(ref_bid)
                    if not ref_drug:
                        continue
                    ref_conc = parse_concentration(ref_drug)
                    if not ref_conc:
                        ref_tl = ref_drug.get("tl") or {}
                        pc = ref_tl.get("pc") or {}
                        per_ml = pc.get("per_ml")
                        raw = pc.get("raw")
                        if isinstance(per_ml, (int, float)) and pc.get("kind") == "mg":
                            ref_conc = float(per_ml), str(raw or "")
                    if not ref_conc or unit_kind != "mg":
                        continue
                    ref_per_ml, _ = ref_conc
                    ml_range = parse_ref_volume_ml(row.get("o"), row.get("cv"))
                    if not ml_range:
                        continue
                    mg_lo = ml_range[0] * ref_per_ml
                    mg_hi = ml_range[1] * ref_per_ml
                    frac_lo = mg_lo / unit_strength
                    frac_hi = mg_hi / unit_strength
                    for frac in (frac_lo, frac_hi):
                        if frac <= 0 or frac > 4:
                            break
                        nearest_quarter = round(frac * 4) / 4
                        if abs(frac - nearest_quarter) > 1e-6:
                            break
                    else:
                        unit_label = "tab" if "tab" in form else "cap"
                        order = f"{format_fraction_units(frac_lo, frac_hi, unit_label)} {str(row.get('f') or '').strip()}".strip()
                        derived = {
                            "rp": row.get("n") or ref_drug.get("n") or "",
                            "rs": order,
                            "rf": str(row.get("f") or "").strip(),
                            "rd": str(row.get("u") or "").strip(),
                            "rdi": "",
                            "nt": "Derived solid oral pediatric fraction from same-generic liquid reference.",
                        }
                        break
                if derived:
                    break
            if derived:
                tl.update(derived)
                tl["s"] = "reference_default_derived_from_same_generic"
                d["tl"] = tl
                annotate_peds_quality(d)
                upgraded_reference_fraction += 1
                continue

        annotate_peds_quality(d)
        if (d.get("tl") or {}).get("q") == "manual_unknown":
            gcur = canonical_generic_key(d.get("g"))
            if any(x in gcur for x in curated_peds_generics):
                d["tl"]["q"] = "parser_manual"
                d["tl"]["nt"] = (d["tl"].get("nt") or "") + " Curated class uplift applied."

    seed.setdefault("m", {})
    seed["m"]["schema_version"] = "druglist-seed-v1"
    seed["m"]["build_version"] = f"auto-{date.today().isoformat()}"
    seed["m"]["price_updated_at"] = latest_date or "unknown"
    seed["m"]["pricedDrugCount"] = sum(1 for d in drugs if isinstance(d.get("pr"), (int, float)) and d.get("pr") > 0)

    OUT.parent.mkdir(exist_ok=True)
    OUT.write_text(json.dumps(seed, ensure_ascii=False, separators=(",", ":")))

    conf = {"direct": 0, "generic_imputed": 0, "category_imputed": 0, "form_imputed": 0, "other": 0}
    for d in drugs:
        c = d.get("price_confidence")
        conf[c if c in conf else "other"] += 1

    pq = {"auto_mapped": 0, "parser_manual": 0, "generic_uplift_manual": 0, "manual_unknown": 0, "none": 0}
    for d in drugs:
        q = (d.get("tl") or {}).get("q", "none")
        pq[q if q in pq else "none"] += 1

    report = {
        "generated_at": date.today().isoformat(),
        "direct_price_updates": direct_price_updates,
        "generic_fallback_updates": generic_fallback_updates,
        "category_fallback_updates": category_fallback_updates,
        "form_fallback_updates": form_fallback_updates,
        "latest_price_date": latest_date,
        "pediatric_upgraded_parser": upgraded_parse,
        "pediatric_upgraded_top_generic": upgraded_top_generic,
        "pediatric_reference_manual_defaults": upgraded_reference_manual,
        "pediatric_reference_fraction_defaults": upgraded_reference_fraction,
        "pediatric_curated_generic_defaults": upgraded_curated_generic,
        "pricedDrugCount": seed["m"].get("pricedDrugCount"),
        "price_confidence_counts": conf,
        "pediatric_quality_counts": pq,
        "drugCount": len(drugs),
    }
    REPORT.write_text(json.dumps(report, indent=2, ensure_ascii=False))
    print(json.dumps(report, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
